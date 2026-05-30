"""Timesheet export to CSV / XLSX (Phase 8).

Turns a `queries.timesheet(...)` result into a downloadable file. CSV uses the
stdlib `csv` module (always available). XLSX needs `openpyxl`, an optional
extra (`pip install -e ".[export]"`) — kept optional so the core install stays
lean and CSV export works everywhere.

Both writers return `bytes`, so the API layer can stream them as a response
without touching the filesystem.
"""

from __future__ import annotations

import csv
import io

_HEADER = ["App", "Screenshots", "Minutes", "Hours"]
_XLSX_INSTALL_HINT = 'XLSX export needs openpyxl. Install with: pip install -e ".[export]"'


def _rows_for_output(timesheet: dict) -> list[list]:
    """Flatten a timesheet dict into header + per-app rows + a total row."""
    out: list[list] = [_HEADER]
    for r in timesheet["rows"]:
        minutes = r["minutes"]
        out.append([r["app_name"], r["shots"], minutes, round(minutes / 60, 2)])
    total_min = timesheet["total_minutes"]
    out.append(["TOTAL", timesheet["total_shots"], total_min, round(total_min / 60, 2)])
    return out


def to_csv(timesheet: dict) -> bytes:
    """Render a timesheet as CSV bytes (UTF-8)."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerows(_rows_for_output(timesheet))
    return buf.getvalue().encode("utf-8")


def to_xlsx(timesheet: dict) -> bytes:
    """Render a timesheet as XLSX bytes. Raises RuntimeError if openpyxl is absent."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as e:
        raise RuntimeError(_XLSX_INSTALL_HINT) from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    rows = _rows_for_output(timesheet)
    for row in rows:
        ws.append(row)

    # Bold the header row and the TOTAL row.
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    for cell in ws[ws.max_row]:
        cell.font = bold

    # Roomy first column for app names.
    ws.column_dimensions["A"].width = 28

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_timesheet(timesheet: dict, fmt: str) -> tuple[bytes, str, str]:
    """Render `timesheet` in `fmt` ('csv'|'xlsx').

    Returns `(data, media_type, file_extension)`. Raises ValueError on an
    unknown format, RuntimeError if XLSX is requested without openpyxl.
    """
    if fmt == "csv":
        return to_csv(timesheet), "text/csv", "csv"
    if fmt == "xlsx":
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return to_xlsx(timesheet), media, "xlsx"
    raise ValueError(f"unsupported export format: {fmt!r} (use 'csv' or 'xlsx')")
